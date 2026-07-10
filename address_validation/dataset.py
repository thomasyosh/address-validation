from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook


@dataclass
class DatasetRow:
    row_id: int
    eaddress: str | None
    caddress: str | None
    easting: float | None
    northing: float | None
    building_csuid: str | None


@dataclass
class FetchTask:
    row_id: int
    address_type: str
    address: str
    easting: float | None
    northing: float | None
    building_csuid: str | None


def load_address_dataset(
    dataset_path: str | Path,
    *,
    id_column: str | None = "id",
    eaddress_column: str = "EADDRESS",
    caddress_column: str = "CADDRESS",
    easting_column: str = "EASTING",
    northing_column: str = "NORTHING",
    building_csuid_column: str | None = "BUILDING_CSUID",
    sheet_name: str | None = None,
) -> list[DatasetRow]:
    path = Path(dataset_path)
    if not path.exists():
        raise FileNotFoundError(f"Dataset not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    rows = worksheet.iter_rows(values_only=True)
    headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    header_index = {header: index for index, header in enumerate(headers)}

    required_columns = [eaddress_column, caddress_column, easting_column, northing_column]
    missing = [column for column in required_columns if column not in header_index]
    if missing:
        raise ValueError(
            f"Dataset must contain columns {required_columns}. Missing: {missing}. Found: {headers}"
        )

    records: list[DatasetRow] = []
    for excel_row_number, row in enumerate(rows, start=2):
        if row is None:
            continue

        row_id = _read_row_id(row, header_index, id_column, excel_row_number)
        eaddress = _read_text(row, header_index, eaddress_column)
        caddress = _read_text(row, header_index, caddress_column)
        if not eaddress and not caddress:
            continue

        records.append(
            DatasetRow(
                row_id=row_id,
                eaddress=eaddress,
                caddress=caddress,
                easting=_read_float(row, header_index, easting_column),
                northing=_read_float(row, header_index, northing_column),
                building_csuid=_read_text(row, header_index, building_csuid_column)
                if building_csuid_column and building_csuid_column in header_index
                else None,
            )
        )

    workbook.close()
    if not records:
        raise ValueError(f"No address rows found in dataset: {path}")

    return records


def iter_fetch_tasks(rows: list[DatasetRow]) -> Iterator[FetchTask]:
    for row in rows:
        if row.eaddress:
            yield FetchTask(
                row_id=row.row_id,
                address_type="EADDRESS",
                address=row.eaddress,
                easting=row.easting,
                northing=row.northing,
                building_csuid=row.building_csuid,
            )
        if row.caddress:
            yield FetchTask(
                row_id=row.row_id,
                address_type="CADDRESS",
                address=row.caddress,
                easting=row.easting,
                northing=row.northing,
                building_csuid=row.building_csuid,
            )


def get_dataset_settings(config: dict[str, Any]) -> dict[str, Any]:
    dataset = config.get("dataset") or {}
    path = dataset.get("path")
    if not path:
        raise ValueError("Config must define dataset.path for routine/benchmark runs.")
    return dataset


def _read_row_id(
    row: tuple[Any, ...],
    header_index: dict[str, int],
    id_column: str | None,
    excel_row_number: int,
) -> int:
    if id_column and id_column in header_index:
        raw_id = row[header_index[id_column]] if header_index[id_column] < len(row) else None
        if raw_id is not None and str(raw_id).strip() != "":
            return int(raw_id)
    return excel_row_number - 1


def _read_text(row: tuple[Any, ...], header_index: dict[str, int], column: str) -> str | None:
    if column not in header_index:
        return None
    index = header_index[column]
    if index >= len(row) or row[index] is None:
        return None
    value = str(row[index]).strip()
    return value or None


def _read_float(row: tuple[Any, ...], header_index: dict[str, int], column: str) -> float | None:
    text = _read_text(row, header_index, column)
    if text is None:
        return None
    try:
        return float(text)
    except ValueError:
        return None
